import flet as ft
from conec_db.contact_inventario_empleado import Contact_Inventario_Empleado
import pandas as pd
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import datetime
import os



class PDF(FPDF):
    def header(self):
        self.set_font('Arial', 'B', 14)
        self.set_fill_color(100, 100, 255)  # Fondo azul para el título
        self.set_text_color(255, 255, 255)  # Texto blanco para el título
        self.cell(0, 10, 'Tabla de Datos', border=0, ln=1, align='C', fill=True)
        self.ln(10)  # Espacio después del título

    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Página {self.page_no()}', 0, 0, 'C')


class Pag_Inventario_Empleado(ft.UserControl):
    def __init__(self, page):
        super().__init__(expand=True)
        self.page = page
        self.data = Contact_Inventario_Empleado()
        self.selected_row = None

        self.nombre = ft.TextField(
            width=280,
            height=40,
            label="Nombre",
            border_color=ft.colors.LIGHT_GREEN_ACCENT_400,
            label_style=ft.TextStyle(color=ft.colors.CYAN_ACCENT_200),
            prefix_icon=ft.icons.PERSON_ADD_ALT_1,
            border='underline'
        )
        self.apellido = ft.TextField(
            width=280,
            height=40,
            label="Apellido",
            border_color=ft.colors.LIGHT_GREEN_ACCENT_400,
            label_style=ft.TextStyle(color=ft.colors.CYAN_ACCENT_200),
            prefix_icon=ft.icons.ARTICLE,
            border='underline'
        )
        self.correo = ft.TextField(
            width=280,
            height=40,
            label="Correo",
            border_color=ft.colors.LIGHT_GREEN_ACCENT_400,
            label_style=ft.TextStyle(color=ft.colors.CYAN_ACCENT_200),
            prefix_icon=ft.icons.EMAIL,
            border='underline'
        )
        self.telefon = ft.TextField(
            width=280,
            height=40,
            label="Telefono",
            border_color=ft.colors.LIGHT_GREEN_ACCENT_400,
            label_style=ft.TextStyle(color=ft.colors.CYAN_ACCENT_200),
            prefix_icon=ft.icons.PHONE,
            border='underline',
            input_filter=ft.NumbersOnlyInputFilter(),
            max_length=9
        )
        self.descripcion = ft.TextField(
            width=280,
            height=40,
            label="Descripcion",
            border_color=ft.colors.LIGHT_GREEN_ACCENT_400,
            label_style=ft.TextStyle(color=ft.colors.CYAN_ACCENT_200),
            prefix_icon=ft.icons.DESCRIPTION,
            border='underline'
        )
        self.imagen = ft.ElevatedButton(
            content=ft.Row([
                ft.Icon(ft.icons.IMAGE, color='white'),
                ft.Text(
                    "Seleccionar imagen",
                    weight='w500',  # Peso del texto
                    color='white',

                )
            ],
                alignment='center'
            ),
            bgcolor='transparent'
        )

        self.searh_field = ft.TextField(
            suffix_icon=ft.icons.SEARCH,
            label="Buscar por el nombre",
            border=ft.InputBorder.UNDERLINE,
            border_color=ft.colors.LIGHT_GREEN_ACCENT_400,
            label_style=ft.TextStyle(color=ft.colors.CYAN_ACCENT_200),
            on_change=self.searh_data,
        )

        self.data_table = ft.DataTable(
            expand=True,
            border=ft.border.all(2, "purple"),
            data_row_color={ft.ControlState.SELECTED: ft.colors.BLUE_600, ft.ControlState.PRESSED: "black"},
            border_radius=10,
            show_checkbox_column=True,
            columns=[
                ft.DataColumn(ft.Text("Nombre", color=ft.colors.LIGHT_GREEN_ACCENT_400, weight="bold")),
                ft.DataColumn(ft.Text("Edad",  color=ft.colors.CYAN_ACCENT_200, weight="bold")),
                ft.DataColumn(ft.Text("Correo",  color=ft.colors.CYAN_ACCENT_200, weight="bold")),
                ft.DataColumn(ft.Text("Telefono",  color=ft.colors.CYAN_ACCENT_200, weight="bold"), numeric=True),
                ft.DataColumn(ft.Text("Descripcion", color=ft.colors.LIGHT_GREEN_ACCENT_400, weight="bold")),
            ],
        )

        self.show_data()

        self.form = ft.Container(
            gradient=ft.LinearGradient(
                [ft.colors.RED, ft.colors.BLUE_ACCENT_400, ft.colors.CYAN_ACCENT_200]),
            border_radius=10,
            col=4,
            padding=10,
            content=ft.Column(
                alignment=ft.MainAxisAlignment.SPACE_AROUND,
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                controls=[
                    ft.Text("Ingrese sus datos",
                            size=40,
                            text_align="center",
                            font_family="vivaldi", ),
                    self.nombre,
                    self.apellido,
                    self.correo,
                    self.telefon,
                    self.descripcion,
                    self.imagen,
                    ft.Container(
                        content=ft.Row(
                            spacing=5,
                            alignment=ft.MainAxisAlignment.SPACE_AROUND,
                            controls=[

                                ft.ElevatedButton(
                                    content=ft.Row(
                                        [
                                            ft.Icon(ft.icons.SAVE, color='white'),  # Icono de guardar
                                            ft.Text(
                                                'Guardar',
                                                weight='w500',  # Peso del texto
                                                color='white',  # Color del texto
                                            ),
                                        ],
                                        alignment='center',
                                    ),
                                    bgcolor='transparent',  # Fondo transparente para el botón
                                    on_click=self.add_data,
                                ),
                                ft.ElevatedButton(
                                    content=ft.Row(
                                        [
                                            ft.Icon(ft.icons.EDIT, color='white'),  # Icono de editar
                                            ft.Text(
                                                'Actualizar',
                                                weight='w500',  # Peso del texto
                                                color='white',  # Color del texto
                                            ),
                                        ],
                                        alignment='center',
                                    ),
                                    bgcolor='transparent',  # Fondo transparente para el botón
                                    on_click=self.update_data,
                                ),
                                ft.ElevatedButton(
                                    content=ft.Row(
                                        [
                                            ft.Icon(ft.icons.DELETE, color='white'),  # Icono de borrar
                                            ft.Text(
                                                'Borrar',
                                                weight='w500',  # Peso del texto
                                                color='white',  # Color del texto
                                            ),
                                        ],
                                        alignment='center',
                                    ),
                                    bgcolor='transparent',  # Fondo transparente para el botón
                                    on_click=self.delete_data,
                                ),
                            ]
                        )
                    )
                ]
            )
        )

        self.table = ft.Container(
            gradient=ft.LinearGradient(
                [ft.colors.RED_ACCENT_200, ft.colors.BLUE_ACCENT_400, ft.colors.CYAN_ACCENT_200]),
            border_radius=10,
            padding=10,
            col=8,
            content=ft.Column(
                expand=True,
                controls=[
                    ft.Container(
                        padding=10,
                        content=ft.Row(
                            controls=[
                                self.searh_field,
                                ft.IconButton(
                                    icon=ft.icons.EDIT,
                                    on_click=self.edit_flied_text,
                                    icon_color="white",
                                ),
                                ft.IconButton(tooltip="Descargar en PDF",
                                              icon=ft.icons.PICTURE_AS_PDF,
                                              icon_color="white",
                                              on_click=self.save_pdf,
                                              ),
                                ft.IconButton(tooltip="Descargar en EXCEL",
                                              icon=ft.icons.SAVE_ALT,
                                              icon_color="white",
                                              on_click=self.save_excel,
                                              ),
                            ]
                        ),
                    ),

                    ft.Column(
                        expand=True,
                        scroll="auto",
                        controls=[
                            ft.ResponsiveRow([
                                self.data_table
                            ]),
                        ]
                    )
                ]
            )
        )
        self.conent = ft.ResponsiveRow(
            controls=[
                self.form,
                self.table
            ]
        )

    def show_data(self):
        self.data_table.rows = []
        for x in self.data.get_contact():
            self.data_table.rows.append(
                ft.DataRow(
                    on_select_changed=self.get_index,
                    cells=[
                        ft.DataCell(ft.Text(x[1])),
                        ft.DataCell(ft.Text(x[2])),
                        ft.DataCell(ft.Text(x[3])),
                        ft.DataCell(ft.Text(str(x[4]))),
                        ft.DataCell(ft.Text(x[5])),
                    ]
                )
            )
        self.update()

    def add_data(self, e):
        nombre = self.nombre.value
        apellido = self.apellido.value
        correo = self.correo.value
        telefon = str(self.telefon.value)
        descripcion = self.descripcion.value

        if len(nombre) and len(apellido) and len(correo) and len(telefon) and len(descripcion) > 0:
            contact_exists = False
            for row in self.data.get_contact():
                if row[1] == nombre:
                    contact_exists = True
                    break

            if not contact_exists:
                self.clean_fields()
                self.data.add_contact(nombre, apellido, correo, telefon, descripcion)
                self.show_data()
            else:
                print("El contacto ya existe en la base de datos.")
        print("Escriba sus datos")

    def get_index(self, e):
        if e.control.selected:
            e.control.selected = False
        else:
            e.control.selected = True
        nombre = e.control.cells[0].content.value
        for row in self.data.get_contact():
            if row[1] == nombre:
                self.selected_row = row
                break
        self.update()

    def edit_flied_text(self, e):
        try:
            self.nombre.value = self.selected_row[1]
            self.apellido.value = self.selected_row[2]
            self.correo.value = self.selected_row[3]
            self.telefon.value = self.selected_row[4]
            self.descripcion.value = self.selected_row[5]
            self.update()
        except TypeError:
            print("Error")

    def update_data(self, e):
        nombre = self.nombre.value
        apellido = self.apellido.value
        correo = self.correo.value
        telefon = str(self.telefon.value)
        descripcion = self.descripcion.value

        if len(nombre) and len(apellido) and len(correo) and len(telefon) and len(descripcion) > 0:
            self.clean_fields()
            self.data.update_contact(self.selected_row[0], nombre, apellido, correo, telefon, descripcion)
            self.show_data()

    def delete_data(self, e):
        self.data.delete_contact(self.selected_row[1])
        self.show_data()

    def searh_data(self, e):
        search = self.searh_field.value.lower()
        nombre = list(filter(lambda x: search in x[1].lower(), self.data.get_contact()))
        self.data_table.rows = []
        if not self.searh_field.value == "":
            if len(nombre) > 0:
                for x in nombre:
                    self.data_table.rows.append(
                        ft.DataRow(
                            on_select_changed=self.get_index,
                            cells=[
                                ft.DataCell(ft.Text(x[1])),
                                ft.DataCell(ft.Text(x[2])),
                                ft.DataCell(ft.Text(x[3])),
                                ft.DataCell(ft.Text(str(x[4]))),
                                ft.DataCell(ft.Text(x[5])),
                            ]
                        )
                    )
                    self.update()
        else:
            self.show_data()

    def clean_fields(self):
        self.nombre.value = ""
        self.apellido.value = ""
        self.correo.value = ""
        self.telefon.value = ""
        self.descripcion.value = ""
        self.update()

    def save_pdf(self, e):
        pdf = PDF()
        pdf.add_page()
        column_widths = [15, 30, 30, 30, 30, 60]  # Ajuste de columnas, mayor ancho para la descripción
        data = self.data.get_contact()
        header = ("ID", "NOMBRE", "APELLIDO", "CORREO", "TELEFONO", "DESCRIPCION")
        data.insert(0, header)

        # Colores para el encabezado
        pdf.set_fill_color(100, 100, 255)  # Fondo azul para el encabezado
        pdf.set_text_color(255, 255, 255)  # Texto blanco para el encabezado
        pdf.set_font('Arial', 'B', 10)

        # Dibujar encabezado de la tabla
        for item, width in zip(header, column_widths):
            pdf.cell(width, 10, item, border=1, align='C', fill=True)
        pdf.ln()

        # Filas de la tabla
        pdf.set_text_color(0, 0, 0)  # Texto negro para las filas
        pdf.set_font('Arial', '', 10)

        # Alternar color de fondo para las filas
        fill = False
        for row in data[1:]:
            pdf.set_fill_color(230, 240, 255) if fill else pdf.set_fill_color(255, 255, 255)
            for item, width in zip(row, column_widths):
                pdf.cell(width, 10, str(item), border=1, align='C', fill=True)
            pdf.ln()
            fill = not fill  # Alternar el color de fondo

        # Guardar el archivo PDF
        file_name = self.get_file_name("pdf")
        pdf.output(file_name)

        # Abrir el PDF automáticamente
        self.open_file(file_name)

    def save_excel(self, e):
        # Crear un nuevo archivo de Excel con openpyxl
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos"

        # Personalización de estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                        bottom=Side(style='thin'))

        # Encabezado de la tabla
        headers = ["ID", "NOMBRE", "APELLIDO", "CORREO", "TELEFONO", "DESCRIPCION"]
        ws.append(headers)

        # Aplicar estilos al encabezado
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # Obtener datos
        data = self.data.get_contact()

        # Agregar filas de datos y aplicar estilos
        for row_num, row in enumerate(data, 2):
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.fill = cell_fill if row_num % 2 == 0 else PatternFill()  # Alternar color de fondo
                cell.border = border
                cell.alignment = Alignment(horizontal="center")

        # Ajustar el ancho de las columnas
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            adjusted_width = max_length + 2
            ws.column_dimensions[col[0].column_letter].width = adjusted_width

        # Guardar el archivo Excel
        file_name = self.get_file_name("xlsx")
        wb.save(file_name)

        # Abrir el archivo de Excel automáticamente
        self.open_file(file_name)

    def get_file_name(self, ext):
        # Obtener la carpeta de documentos del usuario
        documents_folder = os.path.join(os.path.expanduser("~"))

        # Crear la carpeta si no existe
        if not os.path.exists(documents_folder):
            os.makedirs(documents_folder)

        # Generar la ruta completa del archivo
        file_name = os.path.join(documents_folder,
                                 datetime.datetime.now().strftime("DATA_%Y-%m-%d_%H-%M-%S") + f".{ext}")
        return file_name

    def open_file(self, file_name):
        if os.name == 'nt':  # Si es Windows
            os.startfile(file_name)
        elif os.name == 'posix':  # Si es Linux/Mac
            os.system(f'open {file_name}')

    def build(self):
        return self.conent


def main_empleado(page: ft.Page):
    page.bgcolor = "black"
    page.title = "Empleados"
    page.window.min_width = 1100
    page.window.min_height = 500
    form_ui = Pag_Inventario_Empleado(page)
    form_ui.data.close_connection()
    page.add(Pag_Inventario_Empleado(page))


ft.app(main_empleado)